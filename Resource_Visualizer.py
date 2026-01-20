import streamlit as st
import pandas as pd
import plotly.express as px
import plotly.graph_objects as go
import io
import numpy as np
import json
import os
from datetime import datetime
from dateutil.relativedelta import relativedelta

# --------------------------------------------------------------------------
# 0. Library & Settings
# --------------------------------------------------------------------------
st.set_page_config(page_title="Shipyard Planning & Resource Visualizer", layout="wide")

try:
    from openpyxl import Workbook
    from openpyxl.utils.dataframe import dataframe_to_rows
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
except ImportError:
    st.error("❌ 'openpyxl' library is missing. Please run `pip install openpyxl` in your terminal.")
    st.stop()

# --------------------------------------------------------------------------
# [Function] Load Configuration from Text File
# --------------------------------------------------------------------------
def load_config(filename="pre_setting.txt"):
    """
    Load settings from a JSON formatted text file.
    If file doesn't exist or error occurs, return default values.
    """
    default_config = {
        "mh_per_month": 141.6,
        "Assembly": {"type": "front_loaded", "sigma": 0.2},
        "Erection": {"type": "normal", "sigma": 0.15},
        "Painting": {"type": "s_curve", "sigma": 0.25},
        "Outfitting": {"type": "s_curve", "sigma": 0.25},
        "Sea Trial": {"type": "back_loaded", "sigma": 0.2}
    }

    if os.path.exists(filename):
        try:
            with open(filename, "r", encoding='utf-8') as f:
                loaded_config = json.load(f)
                for key, value in loaded_config.items():
                    if key in default_config:
                        default_config[key] = value
        except Exception as e:
            st.warning(f"⚠️ Error loading {filename}: {e}. Using defaults.")
    
    return default_config

# --------------------------------------------------------------------------
# [Function] Detect Overlaps
# --------------------------------------------------------------------------
def detect_overlaps(df):
    overlaps = set()
    messages = []
    
    if not {'Yard', 'Berth', 'KL', 'LC', 'Ship No.'}.issubset(df.columns):
        return overlaps, messages

    for (yard, berth), group in df.groupby(['Yard', 'Berth']):
        ships = group.to_dict('records')
        n = len(ships)
        for i in range(n):
            for j in range(i + 1, n):
                s1 = ships[i]
                s2 = ships[j]
                
                if pd.notnull(s1['KL']) and pd.notnull(s1['LC']) and \
                   pd.notnull(s2['KL']) and pd.notnull(s2['LC']):
                    
                    if (s1['KL'] < s2['LC']) and (s2['KL'] < s1['LC']):
                        overlaps.add(s1['Ship No.'])
                        overlaps.add(s2['Ship No.'])
                        s1_range = f"{s1['KL'].strftime('%Y-%m-%d')} ~ {s1['LC'].strftime('%Y-%m-%d')}"
                        s2_range = f"{s2['KL'].strftime('%Y-%m-%d')} ~ {s2['LC'].strftime('%Y-%m-%d')}"
                        msg = (f"🔴 **[{yard} - {berth}] Overlap Detected**\n"
                               f"   • **{s1['Ship No.']}**: `{s1_range}`\n"
                               f"   • **{s2['Ship No.']}**: `{s2_range}`")
                        messages.append(msg)
    return overlaps, messages

# --------------------------------------------------------------------------
# [Function] Distribute Man-Hours & Logic
# --------------------------------------------------------------------------
def get_pdf_curve(x, curve_type, sigma=0.2):
    if curve_type == 'front_loaded': mean = 0.3
    elif curve_type == 'back_loaded': mean = 0.7
    else: mean = 0.5
    return np.exp(-0.5 * ((x - mean) / sigma) ** 2)

def get_distribution_weights(duration, curve_type='normal', sigma=0.2):
    if duration <= 0: return []
    if duration == 1: return np.array([1.0])
    x = np.linspace(0, 1, duration)
    pdf = get_pdf_curve(x, curve_type, sigma)
    return pdf / np.sum(pdf)

def plot_dynamic_curves(curve_settings):
    x = np.linspace(0, 1, 100)
    fig = go.Figure()
    colors = {'Assembly': '#1f77b4', 'Erection': '#d62728', 'Painting': '#2ca02c', 'Outfitting': '#9467bd', 'Sea Trial': '#17becf'}
    
    for disc, settings in curve_settings.items():
        c_type = settings['type']
        c_sigma = settings['sigma']
        y_vals = get_pdf_curve(x, c_type, c_sigma)
        fig.add_trace(go.Scatter(x=x, y=y_vals, mode='lines', name=f"{disc} ({c_type}, σ={c_sigma})", line=dict(color=colors.get(disc, 'grey'), width=2)))
    
    fig.update_layout(title="Reference: Current Distribution Profiles", xaxis_title="Progress (0~100%)", yaxis_title="Intensity", height=300, margin=dict(l=20,r=20,t=40,b=20))
    return fig

def distribute_mh_for_row(row, curve_settings):
    distribution_result = {} 
    dates = {'SC': row.get('SC'), 'KL': row.get('KL'), 'LC': row.get('LC'), 'DL': row.get('DL')}
    if pd.isnull(dates['SC']) or pd.isnull(dates['DL']): return {} 

    def distribute_discipline(disc_name, mh_col, start_dt, end_dt):
        raw_val = str(row.get(mh_col, 0)).replace(',', '').strip()
        mh_val = pd.to_numeric(raw_val, errors='coerce')
        
        if mh_val > 0 and pd.notnull(start_dt) and pd.notnull(end_dt) and start_dt <= end_dt:
            s_month = start_dt.replace(day=1)
            e_month = end_dt.replace(day=1)
            months = pd.date_range(start=s_month, end=e_month, freq='MS')
            if len(months) == 0: months = [s_month]
            
            setting = curve_settings.get(disc_name)
            weights = get_distribution_weights(len(months), setting['type'], setting['sigma'])
            for i, m in enumerate(months):
                if m not in distribution_result: distribution_result[m] = {}
                distribution_result[m][disc_name] = mh_val * weights[i]

    lc_safe = dates['LC'] if pd.notnull(dates['LC']) else dates['KL']
    distribute_discipline('Assembly', 'Assembly MH', dates['SC'], lc_safe)
    distribute_discipline('Erection', 'Erection MH', dates['KL'], dates['LC'])
    distribute_discipline('Painting', 'Painting MH', dates['SC'], dates['DL'])
    distribute_discipline('Outfitting', 'Outfitting MH', dates['SC'], dates['DL'])
    distribute_discipline('Sea Trial', 'Sea Trial MH', dates['LC'], dates['DL'])
    
    mh_total = pd.to_numeric(str(row.get('Total MH', 0)).replace(',', ''), errors='coerce') or 0
    sum_details = sum([pd.to_numeric(str(row.get(c, 0)).replace(',', ''), errors='coerce') or 0 for c in ['Assembly MH', 'Erection MH', 'Painting MH', 'Outfitting MH', 'Sea Trial MH']])
    
    if sum_details == 0 and mh_total > 0:
         start, end = dates['SC'], dates['DL']
         if start <= end:
             months = pd.date_range(start=start.replace(day=1), end=end.replace(day=1), freq='MS')
             if len(months) == 0: months = [start.replace(day=1)]
             weights = get_distribution_weights(len(months), 'normal', 0.2)
             for i, m in enumerate(months):
                 if m not in distribution_result: distribution_result[m] = {}
                 distribution_result[m]['Unspecified'] = mh_total * weights[i]

    return distribution_result

# --------------------------------------------------------------------------
# [Function] Export Excel Logic (Cached for Performance)
# --------------------------------------------------------------------------
def apply_excel_formatting(ws, final_cols, start_data_col_idx):
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    number_format = '#,##0'
    
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        
    for col_idx, col_name in enumerate(final_cols, 1):
        if isinstance(col_name, datetime) or isinstance(col_name, pd.Timestamp):
            ws.cell(row=1, column=col_idx).value = col_name.strftime('%Y-%m')
            
    for row in ws.iter_rows(min_row=2, min_col=start_data_col_idx): 
        for cell in row:
            cell.number_format = number_format

@st.cache_data(show_spinner=False)
def create_detailed_excel(df, curve_settings, mh_per_month):
    """Report: Detailed MH Breakdown per Ship"""
    all_records = []
    for _, row in df.iterrows():
        ship_dist = distribute_mh_for_row(row, curve_settings)
        disc_map = {}
        for date_val, discs in ship_dist.items():
            for disc_name, mh_val in discs.items():
                if disc_name not in disc_map: disc_map[disc_name] = {}
                disc_map[disc_name][date_val] = mh_val
        
        for disc_name, monthly_data in disc_map.items():
            record = {
                'Yard': row.get('Yard', ''),
                'Ship Class': row.get('Ship Class', ''),
                'Ship No.': row.get('Ship No.', ''),
                'Discipline': disc_name,
                'Total MH': sum(monthly_data.values())
            }
            for m_date, val in monthly_data.items():
                record[m_date] = val 
            all_records.append(record)
            
    if not all_records: return None
    
    res_df = pd.DataFrame(all_records)
    info_cols = ['Yard', 'Ship Class', 'Ship No.', 'Discipline', 'Total MH']
    date_cols = sorted([c for c in res_df.columns if c not in info_cols])
    final_cols = info_cols + date_cols
    res_df = res_df[final_cols].fillna(0)
    
    output = io.BytesIO()
    wb = Workbook()
    ws = wb.active
    ws.title = "Detailed Plan (MH)"
    for r in dataframe_to_rows(res_df, index=False, header=True): ws.append(r)
    apply_excel_formatting(ws, final_cols, 6)
    wb.save(output)
    output.seek(0)
    return output

@st.cache_data(show_spinner=False)
def create_summary_excel(df, curve_settings, mh_per_month):
    """Report: Yard & Discipline Summary (Headcount)"""
    agg_map = {} 
    
    for _, row in df.iterrows():
        yard = row.get('Yard', 'Unknown')
        ship_dist = distribute_mh_for_row(row, curve_settings)
        
        for m_date, discs in ship_dist.items():
            for disc, mh in discs.items():
                key = (yard, disc)
                if key not in agg_map: agg_map[key] = {}
                agg_map[key][m_date] = agg_map[key].get(m_date, 0) + mh
                
    if not agg_map: return None
    
    records = []
    for (yard, disc), monthly_data in agg_map.items():
        rec = {'Yard': yard, 'Discipline': disc}
        total_mh = sum(monthly_data.values())
        rec['Avg Headcount'] = (total_mh / len(monthly_data)) / mh_per_month if len(monthly_data) > 0 else 0
        
        for m_date, mh_val in monthly_data.items():
            rec[m_date] = mh_val / mh_per_month
        records.append(rec)
        
    res_df = pd.DataFrame(records)
    info_cols = ['Yard', 'Discipline', 'Avg Headcount']
    date_cols = sorted([c for c in res_df.columns if c not in info_cols])
    final_cols = info_cols + date_cols
    res_df = res_df[final_cols].fillna(0)
    
    output = io.BytesIO()
    wb = Workbook()
    ws = wb.active
    ws.title = "Yard Summary (Headcount)"
    for r in dataframe_to_rows(res_df, index=False, header=True): ws.append(r)
    apply_excel_formatting(ws, final_cols, 4)
    wb.save(output)
    output.seek(0)
    return output

# --------------------------------------------------------------------------
# 2. Main App Logic
# --------------------------------------------------------------------------
st.title("⚓ Shipyard Planning & Resource Visualizer")

# [Load Config]
config = load_config("pre_setting.txt")

with st.sidebar:
    st.header("⚙️ General Assumptions")
    mh_per_month = st.number_input(
        "Avail. MH / Person / Month", 
        value=float(config.get("mh_per_month", 141.6)), 
        step=0.1, format="%.1f"
    )
    
    st.divider()
    st.header("📉 Workload Distribution Settings")
    st.caption("Select Curve Type & Sigma (Width).")
    
    curve_options = ['front_loaded', 'normal', 's_curve', 'back_loaded']
    curve_settings = {}
    
    def create_curve_ui(label, key_prefix):
        # Load defaults
        disc_conf = config.get(label, {"type": "normal", "sigma": 0.2})
        default_type = disc_conf.get("type", "normal")
        default_sigma = float(disc_conf.get("sigma", 0.2))
        
        try: type_idx = curve_options.index(default_type)
        except ValueError: type_idx = 1
            
        col1, col2 = st.columns([0.65, 0.35])
        with col1: 
            c_type = st.selectbox(f"{label}", curve_options, index=type_idx, key=f"t_{key_prefix}")
        with col2: 
            c_sigma = st.slider(f"σ", 0.1, 0.5, default_sigma, 0.05, key=f"s_{key_prefix}", label_visibility="collapsed")
        return {'type': c_type, 'sigma': c_sigma}

    curve_settings['Assembly'] = create_curve_ui("Assembly", "asm")
    curve_settings['Erection'] = create_curve_ui("Erection", "ere")
    curve_settings['Painting'] = create_curve_ui("Painting", "pnt")
    curve_settings['Outfitting'] = create_curve_ui("Outfitting", "out")
    curve_settings['Sea Trial'] = create_curve_ui("Sea Trial", "sea")

uploaded_file = st.file_uploader("Upload Excel File (.xlsx)", type=["xlsx", "xls"])

if uploaded_file is not None:
    try:
        xls = pd.ExcelFile(uploaded_file)
        if 'Key Event' in xls.sheet_names:
            df = pd.read_excel(uploaded_file, sheet_name='Key Event')
            st.toast("Loaded 'Key Event' sheet.")
        else:
            df = pd.read_excel(uploaded_file)
            st.toast("Loaded first sheet.")
    except Exception as e:
        st.error(f"❌ Failed to read Excel file: {e}")
        st.stop()

    df.columns = df.columns.str.strip()
    required_cols = ['Yard', 'Berth', 'Ship No.', 'SC', 'KL', 'LC', 'DL']
    
    if not set(required_cols).issubset(df.columns):
        st.error(f"❌ Missing required columns: {list(set(required_cols) - set(df.columns))}")
        st.stop()
    
    for col in ['SC', 'KL', 'LC', 'DL']:
        df[col] = pd.to_datetime(df[col], errors='coerce')
    
    df_plot = df.dropna(subset=['SC', 'KL', 'LC', 'DL', 'Ship No.']).copy()
    if 'Berth' in df_plot.columns:
        df_plot = df_plot[df_plot['Berth'].astype(str).str.strip().str.lower() != 'not allocable']

    if df_plot.empty:
        st.warning("⚠️ No valid data found.")
    else:
        overlapping_ships, overlap_msgs = detect_overlaps(df_plot)

        tab1, tab2 = st.tabs(["📅 Berth Schedule", "👷 Resource Analysis"])
        
        with tab1:
            st.subheader("Docking Schedule (KL ~ LC)")
            if overlap_msgs: st.error(f"⚠️ {len(overlap_msgs)} overlaps detected.")
            
            df_chart = df_plot.sort_values(by=['Yard', 'Berth', 'KL'])
            df_chart['Y_Label'] = df_chart['Yard'].astype(str) + " | " + df_chart['Berth'].astype(str)
            fig = px.timeline(df_chart, x_start="KL", x_end="LC", y="Y_Label", color="Yard", text="Ship No.", height=600)
            fig.update_yaxes(autorange="reversed", title="Yard | Berth")
            st.plotly_chart(fig, use_container_width=True)

        with tab2:
            st.subheader("Detailed Workforce Planning by Discipline")
            with st.expander("ℹ️ Reference: Curve Shapes (Live)", expanded=True):
                st.plotly_chart(plot_dynamic_curves(curve_settings), use_container_width=True)

            unique_yards = sorted(df_plot['Yard'].unique())
            yard_options = ["All Yards (Total)"] + list(unique_yards)
            selected_yard_view = st.selectbox("Select Yard View:", yard_options)
            
            target_df = df_plot if selected_yard_view == "All Yards (Total)" else df_plot[df_plot['Yard'] == selected_yard_view]

            agg_data = {} 
            for _, row in target_df.iterrows():
                ship_dist = distribute_mh_for_row(row, curve_settings)
                for m_date, mh_dict in ship_dist.items():
                    if m_date not in agg_data: agg_data[m_date] = {}
                    for disc, val in mh_dict.items():
                        agg_data[m_date][disc] = agg_data[m_date].get(disc, 0) + val
            
            if not agg_data:
                st.warning("No Man-Hour data calculated.")
            else:
                rows = []
                for d, vals in agg_data.items():
                    rows.append({'Date': d, **vals})
                
                res_df = pd.DataFrame(rows).fillna(0).sort_values('Date')
                full_range = pd.date_range(res_df['Date'].min(), res_df['Date'].max(), freq='MS')
                res_df = res_df.set_index('Date').reindex(full_range, fill_value=0).reset_index().rename(columns={'index':'Date'})

                disc_cols = [c for c in res_df.columns if c != 'Date']
                for col in disc_cols: res_df[col] = res_df[col] / mh_per_month
                res_df['Total_Required'] = res_df[disc_cols].sum(axis=1)
                
                fig2 = go.Figure()
                colors = {'Assembly': '#1f77b4', 'Erection': '#d62728', 'Painting': '#2ca02c', 'Outfitting': '#9467bd', 'Sea Trial': '#17becf'}
                for disc in disc_cols:
                    fig2.add_trace(go.Scatter(x=res_df['Date'], y=res_df[disc], mode='lines', stackgroup='skilled', name=disc, line=dict(width=0.5), fillcolor=colors.get(disc, None)))
                fig2.update_layout(title=f"Workforce Breakdown ({selected_yard_view})", xaxis_title="Date", yaxis_title="Headcount", height=500, hovermode="x unified")
                st.plotly_chart(fig2, use_container_width=True)
                
                # On-Demand Export Section
                st.divider()
                st.subheader("📥 Export Reports")
                st.caption("Click the generate button below to prepare Excel files.")

                if st.button("🔄 Generate Excel Reports", type="primary"):
                    with st.spinner("Generating reports... (This may take a few seconds)"):
                        st.session_state['detail_rpt'] = create_detailed_excel(target_df, curve_settings, mh_per_month)
                        st.session_state['summary_rpt'] = create_summary_excel(target_df, curve_settings, mh_per_month)
                        st.success("Reports ready for download!")

                if 'detail_rpt' in st.session_state and st.session_state['detail_rpt']:
                    col_d1, col_d2 = st.columns(2)
                    with col_d1:
                        st.download_button(
                            label="📄 Ship Detail Report (MH)",
                            data=st.session_state['detail_rpt'],
                            file_name=f"Detail_Plan_{selected_yard_view.replace(' ','_')}.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                        )
                    with col_d2:
                        st.download_button(
                            label="📊 Yard Summary (Headcount)",
                            data=st.session_state['summary_rpt'],
                            file_name=f"Yard_Summary_{selected_yard_view.replace(' ','_')}.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                        )