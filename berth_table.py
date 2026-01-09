import streamlit as st
import pandas as pd
import plotly.express as px
import io
from datetime import datetime

# --------------------------------------------------------------------------
# 0. Library & Settings
# --------------------------------------------------------------------------
st.set_page_config(page_title="Berth Plan Visualizer", layout="wide")

try:
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Border, Side, Alignment, Font
    from openpyxl.utils import get_column_letter
except ImportError:
    st.error("❌ 'openpyxl' library is missing. Please run `pip install openpyxl` in your terminal.")
    st.stop()

# --------------------------------------------------------------------------
# [Function] Detect Overlaps (Modified for Better Readability)
# --------------------------------------------------------------------------
def detect_overlaps(df):
    overlaps = set()
    messages = []
    
    # Group by Yard and Berth
    for (yard, berth), group in df.groupby(['Yard', 'Berth']):
        ships = group.to_dict('records')
        n = len(ships)
        for i in range(n):
            for j in range(i + 1, n):
                s1 = ships[i]
                s2 = ships[j]
                
                # Check Overlap (Strictly overlapping time)
                if (s1['KL'] < s2['LC']) and (s2['KL'] < s1['LC']):
                    overlaps.add(s1['Ship No.'])
                    overlaps.add(s2['Ship No.'])
                    
                    # [수정 1] 날짜 사이에 공백 추가 및 포맷팅
                    s1_range = f"{s1['KL'].strftime('%Y-%m-%d')} ~ {s1['LC'].strftime('%Y-%m-%d')}"
                    s2_range = f"{s2['KL'].strftime('%Y-%m-%d')} ~ {s2['LC'].strftime('%Y-%m-%d')}"
                    
                    # [수정 2] 가독성을 위해 줄바꿈(\n)과 글머리 기호 사용
                    # Markdown 문법: **굵게**, `코드블럭(회색배경)`
                    msg = (f"🔴 **[{yard} - {berth}] Overlap Detected**\n"
                           f"   • **{s1['Ship No.']}**: `{s1_range}`\n"
                           f"   • **{s2['Ship No.']}**: `{s2_range}`")
                    messages.append(msg)
                    
    return overlaps, messages
    
# --------------------------------------------------------------------------
# [Function] Create Excel (English Headers + Yard Merge)
# --------------------------------------------------------------------------
def create_excel_gantt(df, overlapping_ships):
    wb = Workbook()
    ws = wb.active
    ws.title = "Berth Plan"

    if df.empty: return wb

    # 1. Calculate Date Range
    min_ts = df['KL'].min()
    max_ts = df['LC'].max()
    if pd.isnull(min_ts) or pd.isnull(max_ts): return wb

    min_date = min_ts.replace(day=1)
    max_date = max_ts
    date_range = pd.date_range(start=min_date, end=max_date, freq='MS')
    
    # -----------------------------------------------------------
    # 2. Setup Headers (Row 1: Year, Row 2: Month)
    # -----------------------------------------------------------
    header_font = Font(bold=True)
    center_align = Alignment(horizontal='center', vertical='center')
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                         top=Side(style='thin'), bottom=Side(style='thin'))
    
    # Set Width for Yard and Berth columns
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 15
    
    # Merge Header Cells
    ws.merge_cells('A1:A2')
    ws.merge_cells('B1:B2')
    
    ws['A1'].value = "Yard"
    ws['B1'].value = "Berth"
    
    for cell in [ws['A1'], ws['B1']]:
        cell.font = header_font
        cell.alignment = center_align
        cell.border = thin_border

    # Freeze Panes (Scroll starts from C3)
    ws.freeze_panes = 'C3'
    
    date_col_map = {}   
    year_col_range = {} 
    
    # 2-1. Month Header (Row 2)
    for i, date_val in enumerate(date_range):
        col_idx = i + 3 # Start from column C (3)
        date_str = date_val.strftime("%Y-%m")
        year = date_val.year
        month = date_val.month
        
        # Display just the number (e.g., "1", "12") for compactness
        cell = ws.cell(row=2, column=col_idx, value=month) 
        cell.alignment = center_align
        cell.font = header_font
        cell.border = thin_border
        
        ws.column_dimensions[get_column_letter(col_idx)].width = 2
        date_col_map[date_str] = col_idx
        
        if year not in year_col_range:
            year_col_range[year] = [col_idx, col_idx]
        else:
            year_col_range[year][1] = col_idx

    # 2-2. Year Header (Row 1)
    for year, (start_col, end_col) in year_col_range.items():
        ws.merge_cells(start_row=1, start_column=start_col, end_row=1, end_column=end_col)
        cell = ws.cell(row=1, column=start_col, value=str(year))
        cell.alignment = center_align
        cell.font = Font(bold=True, size=12)
        cell.border = thin_border
        cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")

    # -----------------------------------------------------------
    # 3. Fill Data (Group by Yard -> Merge)
    # -----------------------------------------------------------
    colors = ['FFCCCC', 'CCFFCC', 'CCCCFF', 'FFFFCC', 'E0E0E0', 'FFDAB9', 'D8BFD8', 'ADD8E6']
    color_map = {}
    current_row = 3
    
    unique_yards = sorted(df['Yard'].unique())
    
    for yard in unique_yards:
        yard_start_row = current_row
        
        # Get data for this Yard
        yard_df = df[df['Yard'] == yard]
        unique_berths = sorted(yard_df['Berth'].unique())

        for berth in unique_berths:
            ws.row_dimensions[current_row].height = 30

            # Write Berth Name (Column B)
            berth_cell = ws.cell(row=current_row, column=2, value=berth)
            berth_cell.font = header_font
            berth_cell.border = thin_border
            berth_cell.alignment = center_align
            
            # Get ships for this Berth
            berth_data = yard_df[yard_df['Berth'] == berth].sort_values(by='KL')
            
            for _, row in berth_data.iterrows():
                try:
                    ship_no = str(row['Ship No.'])
                    ship_class = str(row['Ship Class'])
                    start_date = row['KL']
                    end_date = row['LC']
                    
                    if pd.isnull(start_date) or pd.isnull(end_date): continue

                    start_key = start_date.strftime("%Y-%m")
                    end_key = end_date.strftime("%Y-%m")

                    if start_key in date_col_map and end_key in date_col_map:
                        start_col = date_col_map[start_key]
                        end_col = date_col_map[end_key]
                        
                        if start_col > end_col: continue

                        try:
                            ws.merge_cells(start_row=current_row, start_column=start_col, 
                                           end_row=current_row, end_column=end_col)
                        except: pass 

                        cell = ws.cell(row=current_row, column=start_col)
                        cell.value = ship_no
                        cell.alignment = center_align
                        cell.border = thin_border
                        
                        # Highlight overlaps in Red Bold
                        if ship_no in overlapping_ships:
                            cell.font = Font(bold=True, color="FF0000")
                        else:
                            cell.font = Font(color="000000")
                        
                        # Background Color by Ship Class
                        if ship_class not in color_map:
                            color_map[ship_class] = colors[len(color_map) % len(colors)]
                        cell.fill = PatternFill(start_color=color_map[ship_class], end_color=color_map[ship_class], fill_type="solid")
                except: continue
            
            current_row += 1
        
        # Merge Yard Cells (Column A)
        yard_end_row = current_row - 1
        if yard_end_row >= yard_start_row:
            ws.merge_cells(start_row=yard_start_row, start_column=1, end_row=yard_end_row, end_column=1)
            yard_cell = ws.cell(row=yard_start_row, column=1, value=yard)
            yard_cell.alignment = center_align
            yard_cell.font = header_font
            yard_cell.border = thin_border

    return wb

# --------------------------------------------------------------------------
# 2. Main App Logic
# --------------------------------------------------------------------------
st.title("🚢 Berth Plan Visualizer (Excel Input)")
st.markdown("Upload your Excel file to generate a **Yard/Berth** Gantt chart and formatted Excel report.")

uploaded_file = st.file_uploader("Upload Excel File (.xlsx)", type=["xlsx", "xls"])

if uploaded_file is not None:
    try:
        df = pd.read_excel(uploaded_file)
    except Exception as e:
        st.error(f"❌ Failed to read Excel file: {e}")
        st.stop()

    # Strip whitespace from column names
    df.columns = df.columns.str.strip()
    
    # Check for English column names
    required_cols = ['Yard', 'Berth', 'KL', 'LC', 'Ship No.', 'Ship Class']
    missing_cols = [col for col in required_cols if col not in df.columns]
    
    if missing_cols:
        st.error(f"❌ Missing required columns: {missing_cols}")
        st.info("Please ensure the Excel file has the following headers: Yard, Berth, Ship Class, Ship No., KL, LC")
        st.stop()
    
    # Date Conversion
    df['KL'] = pd.to_datetime(df['KL'], errors='coerce')
    df['LC'] = pd.to_datetime(df['LC'], errors='coerce')
    
    # Filter valid data
    df_plot = df.dropna(subset=['KL', 'LC', 'Yard', 'Berth', 'Ship No.']).copy()
    df_plot = df_plot[df_plot['Berth'].astype(str).str.strip().str.lower() != 'not allocable']

    if df_plot.empty:
        st.warning("⚠️ No valid data found. Please check the Date format (KL, LC).")
    else:
        # Detect Overlaps (New Logic: Exclude 1-day overlaps)
        overlapping_ships, overlap_msgs = detect_overlaps(df_plot)

        if overlap_msgs:
            st.error(f"⚠️ Found {len(overlap_msgs)} schedule overlaps.")
            with st.expander("View Overlap Details", expanded=True): # 바로 보이게 expanded=True 추천
                for msg in overlap_msgs:
                    st.markdown(msg)  # [수정] write -> markdown
                    st.divider()      # [추가] 항목 간 구분선 추가
        else:
            st.success("✅ No overlaps detected.")
        
        # -------------------------------------------------------
        # 1. Interactive Chart
        # -------------------------------------------------------
        st.subheader("1. Interactive Chart Preview")
        
        # Sort: Yard -> Berth -> Date
        df_plot = df_plot.sort_values(by=['Yard', 'Berth', 'KL'])
        
        # Create Y-Axis Label
        df_plot['Y_Label'] = df_plot['Yard'].astype(str) + " | " + df_plot['Berth'].astype(str)
        
        # Modify Title for Overlaps
        df_plot['Title_Disp'] = df_plot.apply(
            lambda x: f"⛔ {x['Ship No.']}" if x['Ship No.'] in overlapping_ships else x['Ship No.'], axis=1
        )

        fig = px.timeline(
            df_plot, x_start="KL", x_end="LC", 
            y="Y_Label",
            color="Ship Class",
            text="Ship No.", 
            hover_data=["Title_Disp"], 
            height=700
        )
        fig.update_yaxes(autorange="reversed", title="Yard | Berth")
        fig.update_xaxes(dtick="M3", tickformat="%Y-%m", showgrid=True)
        st.plotly_chart(fig, use_container_width=True)

        # -------------------------------------------------------
        # 2. Excel Download
        # -------------------------------------------------------
        st.subheader("2. Download Formatted Excel")
        
        try:
            excel_wb = create_excel_gantt(df_plot, overlapping_ships)
            excel_buffer = io.BytesIO()
            excel_wb.save(excel_buffer)
            excel_buffer.seek(0)
            
            st.download_button(
                label="📥 Download Excel Report",
                data=excel_buffer,
                file_name="Berth Table.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
        except Exception as e:
            st.error(f"❌ Error creating Excel file: {e}")

